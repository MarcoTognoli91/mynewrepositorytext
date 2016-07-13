# -*- coding: utf-8 -*-
'''
INFILTRATION: the goal is to evaluate the ACH (air exchange rate) which is calculated by means of the equation: 3,6*Qi/V. 
The volume of the building, V, will be extracted from the excel file using Building_ventilation_data_Extraction() function ;

ATTENTION: First of all modify the Excel data called "Ventilation_Infiltration"

VENTILATION: it is intentional introduction of air from the outdoors into a building; it is further subdivided into natural 
and mechanical ventilation. The goal is to compute the required ventilation flow rate.

LOADS: once ventilation and infiltration flow rates are computed, it is possible to calculate the sensible and latent loads
for the assigned building.
'''
from openpyxl import *
import os
#change your working directory below for accessing the excel file
#os.chdir("C:\Users\Manoj\Dropbox\Building\Assignment")
#load the excel file and get the desired sheet
ExcelFile = load_workbook("Ventilation_Infiltration.xlsx")
Input = ExcelFile.get_sheet_by_name("Input")


def vent_infilt_load(T_set_in,T_out,w_in,w_out,h_in,h_out,season):
    '''This function takes input data with specified unit of measurement from the loaded Excel file and creates a dictionary 
    assigning to each key the corresponding data in Excel;
    '''
    #extraction of input data from Excel file
    BuildingVentilationDataExtraction = {"Volume [m3]":float(Input.columns[1][0].value),
    "Exposed Surface Area [m2]":float(Input.columns[1][1].value),
    "Type of construction [-]":(Input.columns[1][2].value).encode('utf-8'),
    "Design condition [Cooling/Heating]":(Input.columns[1][3].value).encode('utf-8'),
    "Average Stack Height [m]":float(Input.columns[1][4].value),
    "Inside Temperature [°C]":T_set_in,
    "Outdoor Temperature [°C]":T_out,
    "Effective Leakage Area [cm2]":float(Input.columns[1][7].value),
    "Building conditioned floor area [m2]":float(Input.columns[1][8].value),
    "Number of bedrooms [-]":float(Input.columns[1][9].value),
    "Ventilation supply air floor rate [l/s]":float(Input.columns[1][10].value),
    "Ventilation exhaust air floor rate [l/s]":float(Input.columns[1][11].value),
    "Air sensible heat factor_Cs [W/l*s*K]":float(Input.columns[1][12].value),
    "HRV/ERV Sensible effectiveness [-]":float(Input.columns[1][13].value),
    "Balanced ventilation flow rate [l/s]":float(Input.columns[1][14].value),
    "Other balanced ventilation supply flow rate [l/s]":float(Input.columns[1][15].value),
    "HRV/ERV [y/n]":(Input.columns[1][16].value).encode('utf-8'),
    "Air latent heat factor_Cl [W/l*s]":float(Input.columns[1][17].value),
    "Indoor/Outdoor humidity ratio difference [kg/kg]":float(Input.columns[1][18].value),
    "Air total heat factor_Ct [W/l*s]":float(Input.columns[1][19].value),
    "HRV/ERV total  effectiveness [-]":float(Input.columns[1][20].value),
    "Indoor/Outdoor enthalpy difference [kJ/kg]":float(Input.columns[1][21].value)}
    #return BuildingVentilationDataExtraction    

    #BuildingVentilationDataExtraction = Building_ventilation_data_Extraction() #call the function

    building_types_lib = {"Tight":0.7, "Good":1.4,"Average":2.8,"Leaky":5.6,"Very Leaky":10.4} #dictionary of building types to use in the 
                                                                                            #following function

    '''this function takes as input data the dictionary previously created with the Building_ventilation_data_Extraction function
    and computes the infiltration/ventilation airflow rate and the sensible and latent loads of the buiding, returning a dictionary
    made of both intermediate and final results
    '''

    A_ul = building_types_lib.get(BuildingVentilationDataExtraction.get("Type of construction [-]")) #takes the value of the unit leakage area  
    A_es = BuildingVentilationDataExtraction.get("Exposed Surface Area [m2]") #takes the value of the exposed surface area
    

    A_l = A_ul*A_es #calculate the leakage area

    #cycle to get design condition (cooling/heating) and corresponding coefficient to be used later on in IDF definition
    
    while True: #this is to avoid typos
        ans = season
    
        if ans == "Cooling":
            I0=25
            I1=0.38
            I2=0.12
            break
        elif ans == "Heating":
            I0=51
            I1=0.35
            I2=0.23
            break
    
    
    H = BuildingVentilationDataExtraction.get("Average Stack Height [m]") #takes the value of the stack height
    #T_in = BuildingVentilationDataExtraction.get("Inside Temperature [°C]") #takes the value of the inside temperature
    T_in = T_set_in
    #T_out = BuildingVentilationDataExtraction.get("Outdoor Temperature [°C]") #takes the value of the outdoor temperature
    Delta_T = abs(T_in - T_out) #performs indoor/outdoor difference in temperature
    V = BuildingVentilationDataExtraction.get("Volume [m3]") #takes the value of the volume of the building
    Al_flue = BuildingVentilationDataExtraction.get("Effective Leakage Area [cm2]") #takes the value of the effective leakage area
    
    IDF= (I0 + H * abs(Delta_T) * (I1 + I2 * Al_flue / A_l))/1000 #calculates the Infiltration Driving Force
    Qi = A_l * IDF #calculates the infiltartion airflow rate
    ACH = 3.6 * Qi / V #calculates the air exchange rate
    
    A_cf = BuildingVentilationDataExtraction.get("Building conditioned floor area [m2]") #takes the value of building conditioned floor area
    N_br = BuildingVentilationDataExtraction.get("Number of bedrooms [-]") #takes the value of number of bedrooms
    
    Q_v = 0.05 * A_cf + 3.5 * (N_br + 1) #calculates the required ventilation airflow rate
    
    Q_sup = BuildingVentilationDataExtraction.get("Ventilation supply air floor rate [l/s]") #takes the value of ventilation supply airflow rate
    Q_exh = BuildingVentilationDataExtraction.get("Ventilation exhaust air floor rate [l/s]") #takes the value of ventilation exhaust airflow rate
    
    Q_bal= min(Q_sup,Q_exh) 
    
    Q_unbal= max(Q_sup,Q_exh) - Q_bal

    Q_vi = Q_v + max(Q_unbal, Qi + 0.5 * Q_unbal) #compute the overall ventilation/infiltration airflow rate
    
    #Loads calculation
    
    C_s = BuildingVentilationDataExtraction.get("Air sensible heat factor_Cs [W/l*s*K]") #takes the value of the air sensible heat factor 
    eff_s = BuildingVentilationDataExtraction.get("HRV/ERV Sensible effectiveness [-]") #takes the value of HRV/ERV Sensible effectiveness
    Q_bal_hr = BuildingVentilationDataExtraction.get("Balanced ventilation flow rate [l/s]") #takes the value of the balanced ventilation flow rate
    Q_bal_oth = BuildingVentilationDataExtraction.get("Other balanced ventilation supply flow rate [l/s]") #takes the value of other balanced ventilation supply flow rate
     
    
    q_vi_sensible = C_s * (Q_vi + (1 - eff_s) * Q_bal_hr + Q_bal_oth) * Delta_T #compute the sensible load 
    
    #cycle to compute the latent load, according to the presence or not of the HRV/ERV equipment
    while True: #this is to avoid typos
        ans = BuildingVentilationDataExtraction.get("HRV/ERV [y/n]")
        if ans=="n":
            C_l = BuildingVentilationDataExtraction.get("Air latent heat factor_Cl [W/l*s]") #takes the value of the air latent heat factor
            Delta_W = BuildingVentilationDataExtraction.get("Indoor/Outdoor humidity ratio difference [kg/kg]") #takes the value of the indoor/outdoor humidity ratio difference
            Delta_W = w_out-w_in
            q_vi_latent = C_l * (Q_vi + Q_bal_oth) * Delta_W
            break
            
        else:
            C_t = BuildingVentilationDataExtraction.get("Air total heat factor_Ct [W/l*s]") #takes the value of the air total heat factor
            eff_t = BuildingVentilationDataExtraction.get("HRV/ERV total  effectiveness [-]") #takes the value of the HRV/ERV total effectiveness
            #Delta_h = BuildingVentilationDataExtraction.get("Indoor/Outdoor enthalpy difference [kJ/kg]") #takes the value of the in/out enthalpy difference
            Delta_h = abs(h_in-h_out)
            q_vi_total = C_t * (Q_vi + (1 - eff_t) * Q_bal_hr + Q_bal_oth) * Delta_h
            q_vi_latent = q_vi_total - q_vi_sensible
            break
            
    
    RES={"ACH":ACH,
    "IDF":IDF,
    "Qi":Qi,
    "A_l":A_l,
    "Qv":Q_v,
    "Q_bal":Q_bal,
    "Q_unbal":Q_unbal,
    "Q_vi":Q_vi,
    "q_vi_sensible":q_vi_sensible,
    "q_vi_latent":q_vi_latent}
    #RES = vent_infilt_load() #call the function 
    #print (RES)
    #def write_results():
    '''this function takes the results from the vent_infilt_load function and writes them in a dedicated sheet 'Results' in the original Excel file
    '''
    Results = ExcelFile.get_sheet_by_name("Results")
    
    #writing to the results sheet in the excel file
    Results.cell(row=1, column=2).value = RES["A_l"]
    Results.cell(row=2, column=2).value = RES["IDF"]
    Results.cell(row=3, column=2).value = RES["Qi"]
    Results.cell(row=4, column=2).value = RES["ACH"]
    Results.cell(row=5, column=2).value = RES["Qv"]
    Results.cell(row=6, column=2).value = RES["Q_bal"]
    Results.cell(row=7, column=2).value = RES["Q_unbal"]
    Results.cell(row=8, column=2).value = RES["Q_vi"]
    Results.cell(row=9, column=2).value = RES["q_vi_sensible"]
    Results.cell(row=10, column=2).value = RES["q_vi_latent"]
    
    ExcelFile.save("Ventilation_Infiltration.xlsx")
    #returns a dictionary with all the intermediate and final results    
    return RES
#write_results()